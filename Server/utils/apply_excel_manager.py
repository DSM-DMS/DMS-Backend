from openpyxl.styles import Alignment, Color, Font, PatternFill


def get_cells(student):
    number = student.number

    # Get row
    grade = int(number / 1000)
    if grade == 1:
        row_start = 3
    elif grade == 2:
        row_start = 26
    else:
        row_start = 48
    row = row_start + number % 100 - 1

    # Get cols
    class_ = int(number % 1000 / 100)
    number_col = chr(ord('B') + (class_ - 1) * 4)
    name_col = chr(ord(number_col) + 1)
    status_col = chr(ord(name_col) + 1)

    # Get cells
    number_cell = number_col + str(row)
    name_cell = name_col + str(row)
    status_cell = status_col + str(row)

    return number_cell, name_cell, status_cell


def ready_worksheet(ws):
    for row in range(2, 70):
        for col in range(65, 81):
            cell = ws[chr(col) + str(row)]

            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(bold=True)

    ws['B2'] = ws['F2'] = ws['J2'] = ws['N2'] = ws['B25'] = ws['F25'] = ws['J25'] = ws['N25'] = ws['B47'] = ws['F47'] = ws['J47'] = ws['N47'] = '학번'
    ws['C2'] = ws['G2'] = ws['K2'] = ws['O2'] = ws['C25'] = ws['G25'] = ws['K25'] = ws['O25'] = ws['C47'] = ws['G47'] = ws['K47'] = ws['O47'] = '이름'

    for cell1, cell2, cell3 in zip(list(ws.rows)[1], list(ws.rows)[24], list(ws.rows)[46]):
        cell1.fill = cell2.fill = cell3.fill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('C4C4C4'))
