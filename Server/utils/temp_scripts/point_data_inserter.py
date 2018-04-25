# 2018년 4월 24일 MongoDB server 버전 업그레이드 도중 데이터베이스가 유실되어
# 4월 11일까지의 상벌점 정보만 들어 있는 DB에 정보를 채우기 위한 inserter입니다.

from datetime import datetime
from bson.objectid import ObjectId

import re

from openpyxl import load_workbook

from app.models.account import StudentModel
from app.models.point import PointHistoryModel, PointRuleModel


def extract_point_history_data(history):
    date = re.findall('\d\d\d\d-\d\d\-\d\d', history)[0]
    reason = re.findall('\] .+ \(', history)[0][2:-2]
    point = int(re.findall('\d점', history)[0][0])

    return date, reason, point


def insert():
    start_migration = input('벌점 정보 insert 할거임? (Y/N)')

    if start_migration.upper() == 'N':
        return

    filename = input('Filename > ')
    wb = load_workbook(filename)

    date_criteria = input('어느 날짜부터?')

    for ws in wb.worksheets:
        for row in range(2, 90):
            if not ws['A{}'.format(row)].value:
                break

            student_number = int(ws['A{}'.format(row)].value)
            student = StudentModel.objects(number=student_number).first()

            bad_point_histories = ws['F{}'.format(row)].value
            if bad_point_histories:
                b_p_history_list = bad_point_histories.split('\n')

                for history in b_p_history_list:
                    date, reason, point = extract_point_history_data(history)

                    if date <= date_criteria:
                        continue

                    rule = PointRuleModel.objects(name=reason).first()
                    student.point_histories.append(
                        PointHistoryModel(
                            reason=rule.name,
                            point_type=rule.point_type,
                            point=point,
                            time=datetime.now(),
                            id=ObjectId()
                        )
                    )

                    student.bad_point += point

                    if (student.bad_point - 10) // 5 > student.penalty_level and not student.penalty_training_status:
                        student.penalty_level = student.penalty_level + 1
                        student.penalty_training_status = True

                    student.save()
