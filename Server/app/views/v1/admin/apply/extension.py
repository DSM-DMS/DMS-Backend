from openpyxl import Workbook

from flask import Blueprint, make_response, send_from_directory
from flask_restful import Api


from app.views.v1 import BaseResource
from app.views.v1 import admin_only


from app.models.apply import ExtensionApply11Model, ExtensionApply12Model

from utils.excel_style_manager import get_cell_positions_from_student_number, ready_applyment_worksheet

EXTENSION_CLASSES = ['가온실', '나온실', '다온실', '라온실', '3층', '4층', '5층', '2층']

api = Api(Blueprint('admin-extension-api', __name__))
api.prefix = '/admin/extension'


@api.resource('/11')
class Extension11Download(BaseResource):

    @admin_only
    def get(self):
        """
        11시 연장신청 엑셀 다운로드
        """
        wb = Workbook()
        ws = wb.active

        ready_applyment_worksheet(ws)

        for apply in ExtensionApply11Model.objects:
            student = apply.student

            number_cell, name_cell, status_cell = get_cell_positions_from_student_number(student)

            ws[number_cell] = student.number
            ws[name_cell] = student.name
            ws[status_cell] = EXTENSION_CLASSES[apply.class_ - 1]

        filename = '11.xlsx'

        wb.save('{}'.format(filename))
        wb.close()

        resp = make_response(send_from_directory('../', filename))
        resp.headers.extend({'Cache-Control': 'no-cache'})

        return resp


@api.resource('/12')
class Extension12Download(BaseResource):

    @admin_only
    def get(self):
        """
        12시 연장신청 엑셀 다운로드
        """
        wb = Workbook()
        ws = wb.active

        ready_applyment_worksheet(ws)

        for apply in ExtensionApply12Model.objects:
            student = apply.student

            number_cell, name_cell, status_cell = get_cell_positions_from_student_number(student)

            ws[number_cell] = student.number
            ws[name_cell] = student.name
            ws[status_cell] = EXTENSION_CLASSES[apply.class_ - 1]

        filename = '12.xlsx'

        wb.save('{}'.format(filename))
        wb.close()

        resp = make_response(send_from_directory('../', filename))
        resp.headers.extend({'Cache-Control': 'no-cache'})

        return resp
