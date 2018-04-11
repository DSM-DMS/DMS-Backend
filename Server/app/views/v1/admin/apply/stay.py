from openpyxl import Workbook

from flask import Blueprint, make_response, send_from_directory
from flask_restful import Api
from flasgger import swag_from

from app.views.v1 import BaseResource
from app.views.v1 import admin_only

from app.docs.v1.admin.apply.stay import *
from app.models.v2.apply import StudentModel
from app.models.v2.apply import StayApplyModel

from utils.excel_style_manager import get_cell_positions_from_student_number, ready_applyment_worksheet

api = Api(Blueprint('admin-stay-api', __name__))
api.prefix = '/admin'


@api.resource('/stay')
class StayDownload(BaseResource):
    @swag_from(STAY_DOWNLOAD_GET)
    @admin_only
    def get(self):
        """
        잔류신청 엑셀 다운로드
        """
        wb = Workbook()
        ws = wb.active

        ready_applyment_worksheet(ws)

        for student in StudentModel.objects:
            number_cell, name_cell, status_cell = get_cell_positions_from_student_number(student)

            ws[number_cell] = student.number
            ws[name_cell] = student.name

            apply = StayApplyModel.objects(student=student).first()

            if not apply or apply.value == 4:
                status = '잔류'
            elif apply.value == 1:
                status = '금요 귀가'
            elif apply.value == 2:
                status = '토요 귀가'
            elif apply.value == 3:
                status = '토요 귀사'
            else:
                status = '잔류'

            ws[status_cell] = status

        filename = 'stay.xlsx'

        wb.save('{}'.format(filename))
        wb.close()

        resp = make_response(send_from_directory('../', filename))
        resp.headers.extend({'Cache-Control': 'no-cache'})

        return resp
