from openpyxl import Workbook

from flask import Blueprint, make_response, send_from_directory
from flask_restful import Api


from app.views.v1 import BaseResource
from app.views.v1 import admin_only


from app.models.apply import GoingoutApplyModel, StayApplyModel

from utils.excel_style_manager import get_cell_positions_from_student_number, ready_applyment_worksheet

api = Api(Blueprint('admin-goingout-api', __name__))
api.prefix = '/admin'


@api.resource('/goingout')
class GoingoutDownload(BaseResource):

    @admin_only
    def get(self):
        """
        외출신청 엑셀 다운로드
        """
        wb = Workbook()
        ws = wb.active

        ready_applyment_worksheet(ws)

        for apply in GoingoutApplyModel.objects:
            student = apply.student

            number_cell, name_cell, status_cell = get_cell_positions_from_student_number(student)

            stay_apply = StayApplyModel.objects(student=student).first()

            if stay_apply.value < 3:
                ws[number_cell] = None
                ws[name_cell] = None
                continue
            else:
                ws[number_cell] = student.number
                ws[name_cell] = student.name

            if apply.on_saturday and apply.on_sunday:
                status = '토요일, 일요일 외출'
            elif apply.on_saturday:
                status = '토요일 외출'
            elif apply.on_sunday:
                status = '일요일 외출'
            else:
                status = ''
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['H'].width = 20
            ws.column_dimensions['L'].width = 20
            ws.column_dimensions['P'].width = 20

            ws[status_cell] = status

        filename = 'goingout.xlsx'

        wb.save('{}'.format(filename))
        wb.close()

        resp = make_response(send_from_directory('../', filename))
        resp.headers.extend({'Cache-Control': 'no-cache'})

        return resp
