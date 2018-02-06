from binascii import hexlify
from hashlib import pbkdf2_hmac

from openpyxl import Workbook, load_workbook
from uuid import uuid4

from flask import Blueprint, Response, current_app
from flask_restful import Api, abort, request

from app.models.account import AdminModel, SignupWaitingModel, StudentModel
from app.views import BaseResource

from utils.excel_style_manager import ready_uuid_worksheet

api = Api(Blueprint('system-account-management-api', __name__))
api.prefix = '/system'


@api.resource('/account/admin')
class AdminAccount(BaseResource):
    @BaseResource.system_only
    def post(self):
        """
        관리자 계정 생성
        """
        if not request.is_json:
            abort(400)

        id = request.json['id']
        pw = request.json['pw']

        pw = hexlify(pbkdf2_hmac(
            hash_name='sha256',
            password=pw.encode(),
            salt=current_app.secret_key.encode(),
            iterations=100000
        )).decode('utf-8')

        AdminModel(id=id, pw=pw, name='시스템').save()

        return Response('', 201)

    @BaseResource.system_only
    def delete(self):
        """
        관리자 계정 제거
        """
        if not request.is_json:
            abort(400)

        id = request.json['id']
        admin = AdminModel.objects(id=id).first()

        if not admin:
            return Response('', 204)

        admin.delete()

        return Response('', 200)


@api.resource('/uuid-generate/new')
class NewUUIDGeneration(BaseResource):
    @BaseResource.system_only
    def post(self):
        """
        가입되어 있지 않은 학생을 대상으로 UUID Generation
        이미 준비되어 있던 UUID는 제거되므로 주의 필요
        """
        if not request.is_json:
            abort(400)

        student_list = request.json
        workbooks = [Workbook() for _ in range(12)]

        SignupWaitingModel.objects.delete()

        for grade_str, classes in student_list.items():

            for cls_str, students in classes.items():
                wb = workbooks[(int(grade_str) - 1) * 4 + (int(cls_str) - 1)]
                ws = wb.active
                ready_uuid_worksheet(ws)

                for number_str, name in students.items():
                    student_number = int(grade_str + cls_str + number_str)
                    if StudentModel.objects(number=student_number) or not name:
                        # Already signed student, or name is none
                        continue

                    while True:
                        # Generates unduplicated four_figures uuid
                        uuid = str(uuid4())[:4]

                        if not SignupWaitingModel.objects(uuid=uuid):
                            # Generated UUID doesn't exist
                            break

                    SignupWaitingModel(
                        uuid=uuid,
                        name=name,
                        number=student_number
                    ).save()

                    row = str(int(number_str) + 1)

                    ws['A' + row] = student_number
                    ws['B' + row] = name
                    ws['C' + row] = uuid

                wb.save('uuid_{}_{}.xlsx'.format(grade_str, cls_str))

        return Response('', 201)


@api.resource('/uuid-generate/excel-to-db')
class ExcelUUIDToDB(BaseResource):
    def _uuid_excel_save(self):
        for i in range(1, 4):
            for j in range(1, 5):
                wb = load_workbook('uuid_{0}_{1}.xlsx'.format(i, j)).get_sheet_by_name('Sheet1')
                for k in range(2, 23):
                    if wb['A{0}'.format(k)].value:
                        SignupWaitingModel(
                            uuid=wb['C{0}'.format(k)].value,
                            name=wb['A{0}'.format(k)].value,
                            number=int(wb['B{0}'.format(k)].value)
                        ).save()

    @BaseResource.system_only
    def post(self):
        self._uuid_excel_save()
        return Response('', 201)
