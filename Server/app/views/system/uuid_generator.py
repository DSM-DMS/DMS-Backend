from openpyxl import Workbook
from uuid import uuid4

from flask import Blueprint, Response
from flask_restful import Api, abort, request

from app.models.account import SignupWaitingModel, StudentModel
from app.views import BaseResource

from utils.apply_excel_manager import ready_uuid_worksheet

api = Api(Blueprint('system-uuid-generation-api', __name__))
api.prefix = '/system'


@api.resource('/uuid-generate')
class UUIDGeneration(BaseResource):
    def post(self):
        """
        아직 가입되지 않은 학생들이 가입 가능하도록 UUID Generation
        """
        if not request.is_json:
            abort(400)

        student_list = request.json
        # {
        #     '1': {
        #         '1': {
        #             '01': '사람',
        #             '02': '사람2',
        #             ...
        #         },
        #         '2': {
        #
        #         },
        #         '3': {
        #
        #         },
        #         '4': {
        #
        #         }
        #     },
        #     '2' : {
        #         ...
        #     },
        #     '3': {
        #         ...
        #     }
        # }
        workbooks = [Workbook() for _ in range(12)]

        SignupWaitingModel.objects.delete()

        for grade_str, classes in student_list.items():

            for cls_str, students in classes.items():
                wb = workbooks[(int(grade_str) - 1) * 4 + (int(cls_str) - 1)]
                ws = wb.active
                ready_uuid_worksheet(ws)

                for number_str, name in students.items():
                    student_number = int(grade_str + cls_str + number_str)
                    if StudentModel.objects(number=student_number) or not name:
                        # Already signed student, or name is none
                        continue

                    while True:
                        # Generates unduplicated four_figures uuid
                        uuid = str(uuid4())[:4]

                        if not SignupWaitingModel.objects(uuid=uuid):
                            # Generated UUID doesn't exist
                            break

                    SignupWaitingModel(
                        uuid=uuid,
                        name=name,
                        number=student_number
                    ).save()

                    row = str(int(number_str) + 1)

                    ws['A' + row] = student_number
                    ws['B' + row] = name
                    ws['C' + row] = uuid

                wb.save('uuid_{}_{}.xlsx'.format(grade_str, cls_str))

        return Response('', 201)
