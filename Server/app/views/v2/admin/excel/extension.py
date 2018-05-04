from openpyxl import Workbook

from flask import Blueprint, make_response, send_from_directory
from flask_restful import Api
from flasgger import swag_from

from app.docs.v2.admin.excel.extension import *
from app.models.account import AdminModel
from app.models.apply import ExtensionApply11Model, ExtensionApply12Model
from app.views.v2 import BaseResource, auth_required, json_required

from utils.excel_style_manager import get_cell_positions_from_student_number, ready_applyment_worksheet

api = Api(Blueprint(__name__, __name__, url_prefix='/admin/extension'))


class ExtensionExcelDownload(BaseResource):
    def __init__(self, model, filename, extension='.xlsx'):
        self.EXTENSION_CLASSES = '가온실', '나온실', '다온실', '라온실', '3층', '4층', '5층', '2층'
        self.model = model
        self.filename = filename
        self.extension = extension
        self.filename_full = '{}.{}'.format(self.filename, self.extension)

        super(ExtensionExcelDownload, self).__init__()

    def ready_worksheet(self):
        wb = Workbook()
        ws = wb.active

        ready_applyment_worksheet(ws)

        return wb, ws

    def save_excel(self, wb):
        wb.save('{}'.format(self.filename_full))
        wb.close()

    def generate_excel(self):
        wb, ws = self.ready_worksheet()

        for apply in self.model.objects:
            student = apply.student

            number_cell, name_cell, status_cell = get_cell_positions_from_student_number(student)

            ws[number_cell] = student.number
            ws[name_cell] = student.name
            ws[status_cell] = self.EXTENSION_CLASSES[apply.class_ - 1]

        self.save_excel(wb)

    def get_response_object_with_excel_file(self):
        resp = make_response(send_from_directory('../', self.filename_full))
        resp.headers.extend({'Cache-Control': 'no-cache'})

        return resp


@api.resource('/11')
class Extension11ExcelDownload(ExtensionExcelDownload):
    def __init__(self):
        super(Extension11ExcelDownload, self).__init__(ExtensionApply11Model, '11')
        # TODO filename에 문제 있어서 404 나는거임

    @auth_required(AdminModel)
    @swag_from(EXTENSION_11_EXCEL_DOWNLOAD_GET)
    def get(self):
        """
        11시 연장신청 정보 엑셀 다운로드
        """
        self.generate_excel()

        return self.get_response_object_with_excel_file()


@api.resource('/12')
class Extension12ExcelDownload(ExtensionExcelDownload):
    def __init__(self):
        super(Extension12ExcelDownload, self).__init__(ExtensionApply12Model, '12')

    @auth_required(AdminModel)
    @swag_from(EXTENSION_12_EXCEL_DOWNLOAD_GET)
    def get(self):
        """
        12시 연장신청 정보 엑셀 다운로드
        """
        self.generate_excel()

        return self.get_response_object_with_excel_file()
