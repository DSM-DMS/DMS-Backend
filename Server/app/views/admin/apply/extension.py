from openpyxl import Workbook

from flask import Blueprint, make_response, send_from_directory
from flask_restful import Api
from flasgger import swag_from

from app.support.resources import BaseResource
from app.support.view_decorators import admin_only

from app.docs.admin.apply.extension import *
from app.models.account import StudentModel

from utils.excel_style_manager import get_cell_positions_from_student_number, ready_applyment_worksheet

EXTENSION_CLASSES = ['가온실', '나온실', '다온실', '라온실', '3층', '4층', '5층']

api = Api(Blueprint('admin-extension-api', __name__))
api.prefix = '/admin/extension'


@api.resource('/11')
class Extension11Download(BaseResource):
    @swag_from(EXTENSION_DOWNLOAD_GET)
    @admin_only
    def get(self):
        """
        11시 연장신청 엑셀 다운로드
        """
        wb = Workbook()
        ws = wb.active

        ready_applyment_worksheet(ws)

        for student in StudentModel.objects:
            extension_apply = student.extension_apply_11

            if not extension_apply:
                continue

            number_cell, name_cell, status_cell = get_cell_positions_from_student_number(student)

            ws[number_cell] = student.number
            ws[name_cell] = student.name
            ws[status_cell] = EXTENSION_CLASSES[extension_apply.class_ - 1]

        filename = '11.xlsx'

        wb.save('{}'.format(filename))
        wb.close()

        return send_from_directory('../', filename)


@api.resource('/12')
class Extension12Download(BaseResource):
    @swag_from(EXTENSION_DOWNLOAD_GET)
    @admin_only
    def get(self):
        """
        12시 연장신청 엑셀 다운로드
        """
        wb = Workbook()
        ws = wb.active

        ready_applyment_worksheet(ws)

        for student in StudentModel.objects:
            extension_apply = student.extension_apply_12

            if not extension_apply:
                continue

            number_cell, name_cell, status_cell = get_cell_positions_from_student_number(student)

            ws[number_cell] = student.number
            ws[name_cell] = student.name
            ws[status_cell] = EXTENSION_CLASSES[extension_apply.class_ - 1]

        filename = '12.xlsx'

        wb.save('{}'.format(filename))
        wb.close()

        resp = make_response(send_from_directory('../', filename))
        resp.headers.extend({'Cache-Control': 'no-cache'})

        return resp
